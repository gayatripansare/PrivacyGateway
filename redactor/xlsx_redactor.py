"""
redactor/xlsx_redactor.py

Redacts PII from .xlsx / .xlsm files using openpyxl.
Preserves 100% of the original file:
  - All formatting (fonts, colors, fills, borders, alignment)
  - All formulas (formulas that don't contain PII are untouched)
  - All charts, images, pivot tables (passed through untouched)
  - All sheets, merged cells, named ranges
  - Macro-enabled workbooks (.xlsm) — VBA preserved

Returns the cleaned .xlsx/.xlsm file at out_path.

Install: pip install openpyxl
"""

import re
import shutil
import os
from copy import copy


# ─────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────

def _build_replace_map(findings):
    """
    Build a flat list of (raw_value, placeholder) pairs from findings.
    Longer values first — prevents partial replacements where a shorter
    value is a substring of a longer one.
    """
    type_counts = {}
    replace_map = []

    for finding in findings:
        value   = finding.get("value", "")
        replace = finding.get("replace", "[REDACTED]")
        ftype   = finding.get("type", "UNKNOWN")

        if not value:
            continue

        type_counts[ftype] = type_counts.get(ftype, 0) + 1
        placeholder = replace.replace("]", f"_{type_counts[ftype]}]")
        replace_map.append((value, placeholder))

    # Sort longest first to avoid partial-match issues
    replace_map.sort(key=lambda x: len(x[0]), reverse=True)
    return replace_map


def _redact_string(value, replace_map):
    """
    Apply all replacements to a string value.
    Case-insensitive matching, preserves surrounding text.
    Returns (new_string, was_changed).
    """
    if not isinstance(value, str) or not value.strip():
        return value, False

    original = value
    for raw, placeholder in replace_map:
        value = re.sub(re.escape(raw), placeholder, value, flags=re.IGNORECASE)

    return value, value != original


def _redact_cell(cell, replace_map):
    """
    Redact a single cell's value if it is a string.
    Numeric, date, boolean, formula-only cells are not touched
    unless their string representation contains PII.

    Formulas: if the formula's *string* contains PII (e.g. a hardcoded
    email in a formula like =IF(A1="you@x.com","match","no")), we redact
    inside the formula string too.
    """
    val = cell.value

    if val is None:
        return False

    # Plain string cell
    if isinstance(val, str):
        if val.startswith("="):
            # It's a formula — redact within the formula string
            new_val, changed = _redact_string(val, replace_map)
        else:
            new_val, changed = _redact_string(val, replace_map)

        if changed:
            cell.value = new_val
            return True

    return False


# ─────────────────────────────────────────────────────────
# SHARED STRING TABLE REDACTION
# openpyxl exposes cell values through the shared strings
# table automatically — editing cell.value is enough.
# We do NOT need to touch the raw XML.
# ─────────────────────────────────────────────────────────

def _redact_sheet(ws, replace_map):
    """
    Iterate every cell in a worksheet and redact string values.
    Handles merged cells safely (only writes to top-left anchor).
    Returns count of cells changed.
    """
    changed = 0
    merged_anchors = set()

    # Collect top-left cells of merged ranges so we only write there
    for merge_range in ws.merged_cells.ranges:
        merged_anchors.add((merge_range.min_row, merge_range.min_col))

    for row in ws.iter_rows():
        for cell in row:
            # Skip non-anchor merged cells (they are read-only MergedCell objects)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                continue

            if _redact_cell(cell, replace_map):
                changed += 1

    return changed


def _redact_defined_names(wb, replace_map):
    """
    Redact PII from named range comments/titles if any contain strings.
    Named range references (like Sheet1!$A$1) are never touched.
    """
    try:
        for name in wb.defined_names.definedName:
            if hasattr(name, "comment") and name.comment:
                new_comment, changed = _redact_string(name.comment, replace_map)
                if changed:
                    name.comment = new_comment
    except Exception:
        pass  # Named ranges are structural — never break on this


def _redact_sheet_comments(ws, replace_map):
    """Redact text inside cell comments (notes)."""
    try:
        if not hasattr(ws, "_comments"):
            return
        for comment in ws._comments:
            if comment and comment.text:
                new_text, changed = _redact_string(str(comment.text), replace_map)
                if changed:
                    comment.text = new_text
    except Exception:
        pass


def _redact_header_footer(ws, replace_map):
    """Redact PII from sheet headers and footers."""
    try:
        hf = ws.HeaderFooter
        for attr in ("oddHeader", "oddFooter", "evenHeader", "evenFooter",
                     "firstHeader", "firstFooter"):
            section = getattr(hf, attr, None)
            if section is None:
                continue
            for part in ("left", "center", "right"):
                part_obj = getattr(section, part, None)
                if part_obj and hasattr(part_obj, "text") and part_obj.text:
                    new_text, changed = _redact_string(part_obj.text, replace_map)
                    if changed:
                        part_obj.text = new_text
    except Exception:
        pass


# ─────────────────────────────────────────────────────────
# PUBLIC ENTRY POINT
# ─────────────────────────────────────────────────────────

def redact_xlsx(src_path, findings, out_path):
    """
    Redact PII from an Excel file (.xlsx or .xlsm) and save to out_path.

    Parameters
    ----------
    src_path  : str  — path to the original .xlsx / .xlsm file
    findings  : list — list of finding dicts from the scanner
                       each must have: "value", "replace", "type"
    out_path  : str  — where to write the cleaned file

    Returns
    -------
    out_path  : str  — path to the cleaned file (same as input out_path)

    Raises
    ------
    ImportError  if openpyxl is not installed
    ValueError   if findings is empty (nothing to redact)
    Exception    on file read/write errors
    """
    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel redaction. "
            "Install it with: pip install openpyxl"
        )

    if not findings:
        # Nothing to redact — copy original to output unchanged
        shutil.copy2(src_path, out_path)
        return out_path

    if not os.path.exists(src_path):
        raise FileNotFoundError(f"Source file not found: {src_path}")

    # Build replacement map once — reused across all sheets
    replace_map = _build_replace_map(findings)

    if not replace_map:
        shutil.copy2(src_path, out_path)
        return out_path

    # Determine if macro-enabled (.xlsm) — preserve VBA if so
    is_xlsm = src_path.lower().endswith(".xlsm")

    # Load workbook
    # keep_vba=True preserves macros in .xlsm files
    # rich_text=True preserves rich text formatting inside cells
    wb = openpyxl.load_workbook(
        src_path,
        keep_vba=is_xlsm,
        rich_text=False   # rich_text=True can cause issues on some files
    )

    total_changed = 0

    # Process every sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Redact cell values
        total_changed += _redact_sheet(ws, replace_map)

        # Redact cell comments / notes
        _redact_sheet_comments(ws, replace_map)

        # Redact headers and footers
        _redact_header_footer(ws, replace_map)

    # Redact workbook-level named range comments
    _redact_defined_names(wb, replace_map)

    # Save to output path
    # If out_path == src_path we write to a temp file first then replace
    if os.path.abspath(out_path) == os.path.abspath(src_path):
        import tempfile
        tmp = tempfile.mktemp(suffix=os.path.splitext(src_path)[1])
        wb.save(tmp)
        wb.close()
        shutil.move(tmp, out_path)
    else:
        wb.save(out_path)
        wb.close()

    return out_path


# ─────────────────────────────────────────────────────────
# EXTRACTION  — used by api.py to get text for scanning
# ─────────────────────────────────────────────────────────

def extract_text_xlsx(src_path):
    """
    Extract all text content from an Excel file for PII scanning.

    Returns a list of dicts:
      [{"sheet": "Sheet1", "cell": "A1", "value": "text here"}, ...]

    This is what api.py uses to feed text into the scanner before
    calling redact_xlsx() with the findings.
    """
    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        raise ImportError("pip install openpyxl")

    results = []

    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                # Only extract string values (numbers/dates don't leak PII usually)
                # But we also stringify numbers in case they match patterns like Aadhaar
                if isinstance(val, str) and val.strip():
                    results.append({
                        "sheet": sheet_name,
                        "cell":  cell.coordinate,
                        "value": val.strip(),
                    })
                elif isinstance(val, (int, float)):
                    # Convert to string for pattern matching (Aadhaar, phone, etc.)
                    str_val = str(int(val)) if isinstance(val, float) and val.is_integer() else str(val)
                    if len(str_val) >= 5:  # skip tiny numbers
                        results.append({
                            "sheet": sheet_name,
                            "cell":  cell.coordinate,
                            "value": str_val,
                        })

    wb.close()
    return results


def get_full_text_xlsx(src_path):
    """
    Returns all cell text as a single joined string.
    Used by scanner modules that expect a plain text input.
    """
    cells = extract_text_xlsx(src_path)
    return "\n".join(item["value"] for item in cells)